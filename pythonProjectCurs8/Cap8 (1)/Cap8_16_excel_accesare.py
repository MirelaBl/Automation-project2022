"""
816 Module
Modulul openpyxl - exemple accesare
PF - 29.06.2021 v6
"""  #

# instalare modul:      pip install openpyxl
#                       pip install xlwt - pentru fisiere .xls - exemplele sunt pt .xlsx
# instalare modul imagini: pip install pillow
# documentatie : /https://openpyxl.readthedocs.io/en/stable     next urmatoarele pagini

from openpyxl import Workbook, load_workbook
import datetime

wb1 = load_workbook('sample1.xlsx')    # putem deschide un fisier existent

# accesare si schimbare denumire ws
ws1 = wb1.active
print(ws1.title)
ws1.title = 'Maricica'
print(ws1.title)

# prelucrare informatii
for i in range(2,109):
    for j in range(1,5):
        print(ws1.cell(row=i, column=j).value)
    print('*' * 20)

# accesare si modificare valoare camp
print(ws1['A2'].value)
ws1['A2'] = 'Marinescu'
print(ws1['A2'].value)

# atribuire valoare camp
ws1['h1'] = datetime.datetime.now().isoformat()

for i in range(2, 100):
    ws1[ 'g' + str(i)] = "=CONCATENATE(" + 'a' + str(i) + ',' + '" "' + ',' + 'b' + str(i) + ")"

print(ws1['g2'].value)

wb1.save('sample1.xlsx')
wb1.close()
