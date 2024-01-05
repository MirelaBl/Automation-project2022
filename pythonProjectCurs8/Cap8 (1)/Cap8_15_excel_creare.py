"""
815 Module
Modulul openpyxl - exemple creare
PF - 29.06.2021 v6
"""  #

# instalare modul:      pip install openpyxl
#                       pip install xlwt - pentru fisiere .xls - exemplele sunt pt .xlsx
# instalare modul imagini: pip install pillow
# documentatie : /https://openpyxl.readthedocs.io/en/stable     next urmatoarele pagini

from openpyxl import Workbook, load_workbook
import datetime

wb = Workbook()  # Creeaza obiectul care permite lucrul excel in memoria RAM

ws = wb.active  # Lucreaza in worksheet-ul curent (ws) - il creeaza daca nu exista

ws.title = 'Acum'  # atribuie titlu

ws1 = wb.create_sheet()  # Daca nu specificam numele implicit sheet, sheet1, etc
ws2 = wb.create_sheet('Paul', 1)  # Atribuie numele si pozitia in workbook

ws1.title = 'Ioana'  # Schimba numele sheet-ului

print(wb.sheetnames)  # Printeaza o lista cu foile existente

for sheet in wb:
    print(sheet.title)  # Printeaza foile existente

ws['A1'] = 555  # Populare camp direct. Daca exista il rescrie
print(ws['a1'].value)  # acceseaza valoarea

ws.append((7, 8, 9))  # Adaugare pe primele 3 coloane primul rand complet liber

for i in ['a', 'b', 'c']:
    print(ws[i + '2'].value)

var = ws['C1']  # atribuie o variabila corespunzatoare campului
print(var.value)

var1 = ws.cell(row=6, column=4, value=133)   # Variabila corespunzatoare campului cu asignare
print(var1.value)
print(ws['D6'].value)

var.value = 101  # Modificare/atribuire valoare camp
print(var.value)  # printeaza valoarea celulei

ws.append(["17", 28, 39, 45])

for i in ['a', 'b', 'c', 'd']:
    print(ws[i + '7'].value)
ws.append({"b":'a', 3:'x', 6:'alina', 11: 'Feb, 6 2020'})

for i in ['k', 'b', 'c', 'f']:
    print(ws[i + '8'].value)

ws['h1'] = datetime.datetime.now()	# Populeaza data si ora, transforma automat in formatul excel
ws['h2'] = datetime.datetime.now().strftime('%d.%m.%Y')
wb.save("sample.xlsx")   # Salvarea propriu-zisa a fisierului. Daca exista il rescrie. Poate fi de tip .xlsx sau .xlsm

print(ws['A1'].number_format)      # tip de date numeric

ws['a18'] = 302
ws['a17'] = 417

ws["A21"] = "=SUM(a17:a19)"     # lucru cu formule
var1.value = 1101
ws['A6'] = "=SUM({0},{1})".format(var.value, var1.value)
ws['a24'] = "=CONCATENATE(b8,c8)"

wb.save("sample.xlsx")

data = datetime.datetime.now()

wb.save("{0}_sample.xlsx".format(str(data.year)+'{:02d}'.format(data.month)+str(data.day)))

wb.close()
